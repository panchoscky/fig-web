#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generar_torneo.py — último eslabón del pipeline semanal del Torneo Portafolio 2026.

Lee la hoja `ranking_ordenado` del Excel oficial del torneo (la que produce
build_outputs.py) + el Excel de inscripciones (columna LinkedIn) y escribe
`datos/torneo.json` con el esquema documentado en `datos/torneo.json.ejemplo`.

Al existir datos/torneo.json, torneo/index.html abandona el modo DEMO solo.

MANTIENE EL HISTORIAL: si ya existe un datos/torneo.json de una semana
anterior, conserva el campo `historial` de cada equipo y le agrega la semana
nueva. Así la página puede dibujar la trayectoria (rendimiento pasado) en el
overlay de cada equipo y en las tarjetas PNG/HTML descargables.

Uso:
    python3 generar_torneo.py --excel ranking.xlsx --semana 8 --corte "03 · JUL · 2026"
    python3 generar_torneo.py --excel ranking.xlsx --inscripciones inscripciones.xlsx --semana 8 --corte "03 · JUL · 2026"
    python3 generar_torneo.py --demo          # genera un torneo.json de prueba con datos ficticios

Requiere: openpyxl (pip install openpyxl). El modo --demo no requiere nada.

COLUMNAS ESPERADAS en la hoja ranking_ordenado (nombres flexibles, se buscan
por coincidencia, sin distinguir mayúsculas/tildes):
    equipo/nombre, posicion/pos/rank, puntos/score/total,
    ir, exceso/exc, sharpe, var/var95, mdd/drawdown,
    pts_ir, pts_exc, pts_sharpe, pts_var, pts_mdd, ret_rel/retorno_relativo
Si alguna columna no aparece, se avisa y el campo queda en null — la página
lo muestra como "—" sin romperse. Ajustar ALIAS abajo si el Excel cambia.
"""
import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

RAIZ = Path(__file__).resolve().parent
SALIDA = RAIZ / "datos" / "torneo.json"

# alias de columnas: clave del esquema -> posibles encabezados en el Excel
ALIAS = {
    "nombre":   ["equipo", "nombre", "team", "nombre_equipo"],
    "posicion": ["posicion", "pos", "rank", "ranking", "lugar"],
    "puntos":   ["puntos", "score", "total", "puntaje", "puntos_totales"],
    "retRel":   ["ret_rel", "retrel", "retorno_relativo", "ret_relativo", "exceso_acum"],
    "ir":       ["ir", "information_ratio", "info_ratio"],
    "exc":      ["exc", "exceso", "ret_exceso", "excess", "exceso_anual"],
    "sharpe":   ["sharpe", "ratio_sharpe", "sharpe_ratio"],
    "var95":    ["var95", "var", "var_95", "var95_expost"],
    "mdd":      ["mdd", "drawdown", "max_drawdown", "maximum_drawdown"],
    "pts_ir":     ["pts_ir", "puntos_ir", "p_ir"],
    "pts_exc":    ["pts_exc", "puntos_exc", "p_exc", "pts_exceso"],
    "pts_sharpe": ["pts_sharpe", "puntos_sharpe", "p_sharpe"],
    "pts_var95":  ["pts_var", "pts_var95", "puntos_var", "p_var"],
    "pts_mdd":    ["pts_mdd", "puntos_mdd", "p_mdd"],
}

# columnas del Excel de inscripciones
ALIAS_INSC = {
    "equipo":   ["equipo", "nombre_equipo", "team"],
    "nombre":   ["nombre", "integrante", "nombre_completo", "participante"],
    "linkedin": ["linkedin", "url_linkedin", "perfil_linkedin", "in"],
}


def normalizar(texto):
    """minúsculas, sin tildes, sin espacios: para comparar encabezados."""
    t = unicodedata.normalize("NFD", str(texto or ""))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    return re.sub(r"[\s\.\-]+", "_", t.strip().lower())


def slug(nombre):
    t = normalizar(nombre).replace("_", "-")
    t = re.sub(r"[^a-z0-9\-]", "", t)
    return re.sub(r"-+", "-", t).strip("-") or "equipo"


def mapear_columnas(encabezados, alias):
    """encabezados de la hoja -> índice de columna por clave del esquema."""
    norm = [normalizar(h) for h in encabezados]
    mapa = {}
    for clave, candidatos in alias.items():
        for cand in candidatos:
            if cand in norm:
                mapa[clave] = norm.index(cand)
                break
    return mapa


def leer_hoja(ruta, nombre_hoja=None):
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("Falta openpyxl: pip install openpyxl")
    wb = load_workbook(ruta, data_only=True, read_only=True)
    if nombre_hoja:
        hojas = {normalizar(h): h for h in wb.sheetnames}
        real = hojas.get(normalizar(nombre_hoja))
        if not real:
            sys.exit(f"No existe la hoja '{nombre_hoja}' en {ruta}. Hojas: {wb.sheetnames}")
        ws = wb[real]
    else:
        ws = wb.active
    filas = [[c for c in fila] for fila in ws.iter_rows(values_only=True)]
    wb.close()
    return [f for f in filas if any(v is not None and str(v).strip() for v in f)]


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", ".").replace("%", ""))
    except ValueError:
        return None


def leer_ranking(ruta):
    filas = leer_hoja(ruta, "ranking_ordenado")
    mapa = mapear_columnas(filas[0], ALIAS)
    faltan = [k for k in ("nombre", "posicion", "puntos") if k not in mapa]
    if faltan:
        sys.exit(f"Columnas obligatorias no encontradas en ranking_ordenado: {faltan}\n"
                 f"Encabezados vistos: {filas[0]}\nAjusta ALIAS en generar_torneo.py.")
    avisos = [k for k in ALIAS if k not in mapa]
    if avisos:
        print(f"  AVISO: columnas no encontradas (quedarán en null): {avisos}")

    equipos = []
    for fila in filas[1:]:
        def val(k):
            return fila[mapa[k]] if k in mapa and mapa[k] < len(fila) else None
        nombre = val("nombre")
        if not nombre:
            continue
        equipos.append({
            "id": slug(nombre),
            "nombre": str(nombre).strip(),
            "posicion": int(num(val("posicion")) or 0),
            "puntos": round(num(val("puntos")) or 0.0, 2),
            "retRel": num(val("retRel")),
            "delta": 0,  # se calcula después contra la semana anterior
            "metricas": {k: num(val(k)) for k in ("ir", "exc", "sharpe", "var95", "mdd")},
            "puntosDetalle": {
                "ir": num(val("pts_ir")) or 0, "exc": num(val("pts_exc")) or 0,
                "sharpe": num(val("pts_sharpe")) or 0, "var95": num(val("pts_var95")) or 0,
                "mdd": num(val("pts_mdd")) or 0,
            },
            "miembros": [],
        })
    return equipos


def leer_inscripciones(ruta):
    """equipo (slug) -> [{nombre, linkedin}]"""
    filas = leer_hoja(ruta)
    mapa = mapear_columnas(filas[0], ALIAS_INSC)
    if "equipo" not in mapa or "nombre" not in mapa:
        print(f"  AVISO: inscripciones sin columnas equipo/nombre reconocibles "
              f"({filas[0]}) — miembros quedarán vacíos.")
        return {}
    miembros = {}
    for fila in filas[1:]:
        def val(k):
            return fila[mapa[k]] if k in mapa and mapa[k] < len(fila) else None
        eq, nom = val("equipo"), val("nombre")
        if not eq or not nom:
            continue
        li = str(val("linkedin") or "").strip()
        if li and not li.startswith("http"):
            li = "https://" + li.lstrip("/")
        miembros.setdefault(slug(eq), []).append({"nombre": str(nom).strip(), "linkedin": li})
    return miembros


def cargar_anterior():
    if SALIDA.exists():
        try:
            return json.loads(SALIDA.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError) as e:
            print(f"  AVISO: no pude leer el torneo.json anterior ({e}) — historial parte de cero.")
    return None


def integrar_historial(equipos, anterior, semana):
    """conserva el historial previo por id y agrega la semana actual; calcula delta."""
    previos = {}
    if anterior:
        for eq in anterior.get("equipos", []):
            previos[eq["id"]] = eq
    for eq in equipos:
        prev = previos.get(eq["id"])
        hist = list(prev.get("historial", [])) if prev else []
        hist = [h for h in hist if h.get("semana") != semana]  # re-ejecución de la misma semana
        hist.append({"semana": semana, "puntos": eq["puntos"], "posicion": eq["posicion"]})
        hist.sort(key=lambda h: h["semana"])
        eq["historial"] = hist
        if prev and len(hist) >= 2:
            eq["delta"] = hist[-2]["posicion"] - eq["posicion"]  # + = sube
        if prev and not eq["miembros"]:
            eq["miembros"] = prev.get("miembros", [])


def demo():
    """genera un torneo.json de prueba (mismos 8 equipos del modo DEMO de la página)."""
    import random
    random.seed(2026)
    nombres = ["Compass Advisors", "Alpha Capital", "Bull & Bear", "FENIX Group",
               "Horizon Capital", "Delta Global", "Equinox Wealth", "Granite Partners"]
    equipos = []
    for i, n in enumerate(nombres):
        base = 100 - i * 12.5
        hist = [{"semana": s + 1,
                 "puntos": round(max(0, base - (7 - s) * 4 + random.uniform(-3, 3)), 2),
                 "posicion": i + 1} for s in range(8)]
        equipos.append({
            "id": slug(n), "nombre": n, "posicion": i + 1,
            "puntos": hist[-1]["puntos"], "retRel": round(0.08 - i * 0.01, 4), "delta": 0,
            "metricas": {"ir": round(1 - i * .12, 2), "exc": round(.08 - i * .01, 4),
                         "sharpe": round(1.6 - i * .15, 2), "var95": round(-.01 - i * .002, 4),
                         "mdd": round(.05 + i * .017, 3)},
            "puntosDetalle": {"ir": round(30 - i * 3.7, 2), "exc": round(25 - i * 3.1, 2),
                              "sharpe": round(15 - i * 1.9, 2), "var95": round(15 - i * 1.9, 2),
                              "mdd": round(15 - i * 1.9, 2)},
            "miembros": [], "historial": hist,
        })
    return {"semana": 8, "corte": "03 · JUL · 2026", "equipos": equipos}


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    ap.add_argument("--excel", help="Excel oficial con hoja ranking_ordenado")
    ap.add_argument("--inscripciones", help="Excel de inscripciones (columna LinkedIn)")
    ap.add_argument("--semana", type=int, help="número de semana del corte (1-25)")
    ap.add_argument("--corte", help='fecha del corte, ej: "03 · JUL · 2026"')
    ap.add_argument("--demo", action="store_true", help="escribe un torneo.json ficticio de prueba")
    ap.add_argument("--salida", default=str(SALIDA), help=f"ruta de salida (default {SALIDA})")
    args = ap.parse_args()

    salida = Path(args.salida)
    if args.demo:
        datos = demo()
    else:
        if not (args.excel and args.semana and args.corte):
            ap.error("--excel, --semana y --corte son obligatorios (o usa --demo)")
        equipos = leer_ranking(args.excel)
        if args.inscripciones:
            insc = leer_inscripciones(args.inscripciones)
            for eq in equipos:
                eq["miembros"] = insc.get(eq["id"], [])
        integrar_historial(equipos, cargar_anterior(), args.semana)
        datos = {"semana": args.semana, "corte": args.corte, "equipos": equipos}

    salida.parent.mkdir(parents=True, exist_ok=True)
    salida.write_text(json.dumps(datos, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"OK · {salida} · semana {datos['semana']} · {len(datos['equipos'])} equipos")
    print("La página torneo/index.html saldrá del modo DEMO automáticamente.")


if __name__ == "__main__":
    main()
