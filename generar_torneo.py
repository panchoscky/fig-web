#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generar_torneo.py — último eslabón del pipeline semanal del Torneo Portafolio 2026.

Lee la hoja `ranking_ordenado` del Excel oficial del torneo (la que produce
build_outputs.py) + el Excel de inscripciones (columna LinkedIn) y escribe
`datos/torneo.json` con el esquema documentado en `datos/torneo.json.ejemplo`.

Al existir datos/torneo.json, torneo/index.html abandona el modo DEMO solo.

MANTIENE EL HISTORIAL: si ya existe un datos/torneo.json de una semana
anterior, conserva el campo `historial` de cada equipo y le agrega la semana
nueva. Así la página puede dibujar la trayectoria (rendimiento pasado) en el
overlay de cada equipo y en las tarjetas PNG/HTML descargables.

Uso:
    python3 generar_torneo.py --excel ranking.xlsx --semana 8 --corte "03 · JUL · 2026"
    python3 generar_torneo.py --excel ranking.xlsx --inscripciones inscripciones.xlsx --semana 8 --corte "03 · JUL · 2026"
    python3 generar_torneo.py --demo          # genera un torneo.json de prueba con datos ficticios

Requiere: openpyxl (pip install openpyxl). El modo --demo no requiere nada.

COLUMNAS ESPERADAS en la hoja ranking_ordenado (nombres flexibles, se buscan
por coincidencia, sin distinguir mayúsculas/tildes):
    equipo/nombre, posicion/pos/rank, puntos/score/total,
    ir, exceso/exc, sharpe, var/var95, mdd/drawdown,
    pts_ir, pts_exc, pts_sharpe, pts_var, pts_mdd, ret_rel/retorno_relativo
Si alguna columna no aparece, se avisa y el campo queda en null — la página
lo muestra como "—" sin romperse. Ajustar ALIAS abajo si el Excel cambia.
"""
import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

RAIZ = Path(__file__).resolve().parent
SALIDA = RAIZ / "datos" / "torneo.json"

# alias de columnas: clave del esquema -> posibles encabezados en el Excel
ALIAS = {
    "nombre":   ["equipo", "nombre", "team", "nombre_equipo"],
    "posicion": ["posicion", "pos", "rank", "ranking", "lugar"],
    "puntos":   ["puntos", "score", "total", "puntaje", "puntos_totales", "puntaje_total"],
    "retRel":   ["ret_rel", "retrel", "retorno_relativo", "ret_relativo", "exceso_acum"],
    "ir":       ["ir", "information_ratio", "info_ratio"],
    "exc":      ["exc", "exceso", "ret_exceso", "excess", "exceso_anual"],
    "sharpe":   ["sharpe", "ratio_sharpe", "sharpe_ratio"],
    "var95":    ["var95", "var", "var_95", "var95_expost"],
    "mdd":      ["mdd", "drawdown", "max_drawdown", "maximum_drawdown"],
    "pts_ir":     ["pts_ir", "puntos_ir", "p_ir"],
    "pts_exc":    ["pts_exc", "puntos_exc", "p_exc", "pts_exceso"],
    "pts_sharpe": ["pts_sharpe", "puntos_sharpe", "p_sharpe"],
    "pts_var95":  ["pts_var", "pts_var95", "puntos_var", "p_var"],
    "pts_mdd":    ["pts_mdd", "puntos_mdd", "p_mdd"],
    "ret":        ["ret_acum", "retorno_acum", "ret_total", "retorno_acumulado", "retorno"],
}

# columnas del Excel de inscripciones (formato largo: una fila por integrante)
ALIAS_INSC = {
    "equipo":   ["equipo", "nombre_equipo", "team"],
    "nombre":   ["nombre", "integrante", "nombre_completo", "participante"],
    "linkedin": ["linkedin", "url_linkedin", "perfil_linkedin", "in"],
}

# columnas de la hoja "Tabla" / "puntos" del Excel oficial: métricas y puntaje
# por métrica, más completas que ranking_ordenado (que solo trae el total).
ALIAS_TABLA = {
    "nombre": ["equipo"],
    "ir": ["ir"],
    "exc": ["ret_exceso", "ret_exceso_%", "exceso"],
    "sharpe": ["sharpe"],
    "var95": ["var_95", "var_95_%", "var95"],
    "mdd": ["mdd", "mdd_%"],
}
ALIAS_PUNTOS = {
    "nombre": ["equipo"],
    "pts_ir": ["pts_ir"],
    "pts_exc": ["pts_retexc", "pts_exc"],
    "pts_sharpe": ["pts_sharpe"],
    "pts_var95": ["pts_var"],
    "pts_mdd": ["pts_mdd"],
}


def normalizar(texto):
    """minúsculas, sin tildes, sin espacios ni símbolos: para comparar encabezados."""
    t = unicodedata.normalize("NFD", str(texto or ""))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    t = re.sub(r"[\s\.\-]+", "_", t.strip().lower())
    t = re.sub(r"[^a-z0-9_]", "", t)
    return re.sub(r"_+", "_", t).strip("_")


def slug(nombre):
    t = normalizar(nombre).replace("_", "-")
    t = re.sub(r"[^a-z0-9\-]", "", t)
    return re.sub(r"-+", "-", t).strip("-") or "equipo"


def mapear_columnas(encabezados, alias):
    """encabezados de la hoja -> índice de columna por clave del esquema."""
    norm = [normalizar(h) for h in encabezados]
    mapa = {}
    for clave, candidatos in alias.items():
        for cand in candidatos:
            if cand in norm:
                mapa[clave] = norm.index(cand)
                break
    return mapa


def leer_hoja(ruta, nombre_hoja=None, opcional=False):
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("Falta openpyxl: pip install openpyxl")
    wb = load_workbook(ruta, data_only=True, read_only=True)
    if nombre_hoja:
        hojas = {normalizar(h): h for h in wb.sheetnames}
        real = hojas.get(normalizar(nombre_hoja))
        if not real:
            if opcional:
                wb.close()
                return None
            sys.exit(f"No existe la hoja '{nombre_hoja}' en {ruta}. Hojas: {wb.sheetnames}")
        ws = wb[real]
    else:
        ws = wb.active
    filas = [[c for c in fila] for fila in ws.iter_rows(values_only=True)]
    wb.close()
    return [f for f in filas if any(v is not None and str(v).strip() for v in f)]


def leer_hoja_con_encabezado(ruta, nombre_hoja, alias, opcional=True):
    """Como leer_hoja, pero busca la fila de encabezado (puede no ser la 1a,
    ej. hojas con título arriba) buscando la primera fila que reconozca la
    columna 'nombre' del alias dado. Devuelve (filas_de_datos, mapa) o
    (None, None) si la hoja no existe o no se reconoce el encabezado."""
    filas = leer_hoja(ruta, nombre_hoja, opcional=opcional)
    if not filas:
        return None, None
    for i, fila in enumerate(filas[:10]):
        mapa = mapear_columnas(fila, alias)
        if "nombre" in mapa:
            return filas[i + 1:], mapa
    return None, None


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", ".").replace("%", ""))
    except ValueError:
        return None


def leer_metricas_y_puntos(ruta):
    """Lee las hojas 'Tabla' (métricas crudas) y 'puntos' (puntaje por
    métrica) del Excel oficial, más completas que ranking_ordenado (que solo
    trae el total). Devuelve (metricas_por_slug, puntosDetalle_por_slug)."""
    metricas, puntos_detalle = {}, {}

    filas, mapa = leer_hoja_con_encabezado(ruta, "Tabla", ALIAS_TABLA)
    if filas:
        for fila in filas:
            def val(k, f=fila, m=mapa):
                return f[m[k]] if k in m and m[k] < len(f) else None
            nombre = val("nombre")
            if not nombre:
                continue
            metricas[slug(nombre)] = {
                "ir": num(val("ir")),
                "exc": (num(val("exc")) or 0) / 100,
                "sharpe": num(val("sharpe")),
                "var95": (num(val("var95")) or 0) / 100,
                "mdd": (num(val("mdd")) or 0) / 100,
            }

    filas, mapa = leer_hoja_con_encabezado(ruta, "puntos", ALIAS_PUNTOS)
    if filas:
        for fila in filas:
            def val(k, f=fila, m=mapa):
                return f[m[k]] if k in m and m[k] < len(f) else None
            nombre = val("nombre")
            if not nombre:
                continue
            puntos_detalle[slug(nombre)] = {
                "ir": num(val("pts_ir")) or 0, "exc": num(val("pts_exc")) or 0,
                "sharpe": num(val("pts_sharpe")) or 0, "var95": num(val("pts_var95")) or 0,
                "mdd": num(val("pts_mdd")) or 0,
            }
    return metricas, puntos_detalle


def leer_ranking(ruta):
    filas = leer_hoja(ruta, "ranking_ordenado")
    mapa = mapear_columnas(filas[0], ALIAS)
    faltan = [k for k in ("nombre", "posicion", "puntos") if k not in mapa]
    if faltan:
        sys.exit(f"Columnas obligatorias no encontradas en ranking_ordenado: {faltan}\n"
                 f"Encabezados vistos: {filas[0]}\nAjusta ALIAS en generar_torneo.py.")
    avisos = [k for k in ALIAS if k not in mapa]
    if avisos:
        print(f"  AVISO: columnas no encontradas en ranking_ordenado (se completan "
              f"con Tabla/puntos si existen): {avisos}")

    metricas_extra, puntos_extra = leer_metricas_y_puntos(ruta)

    equipos = []
    for fila in filas[1:]:
        def val(k):
            return fila[mapa[k]] if k in mapa and mapa[k] < len(fila) else None
        nombre = val("nombre")
        if not nombre or normalizar(nombre) in ("consultar", "por_consultar"):
            continue  # fila placeholder del Excel, no es un equipo real
        eq_slug = slug(nombre)
        metricas = {k: num(val(k)) for k in ("ir", "exc", "sharpe", "var95", "mdd")}
        if not any(metricas.values()) and eq_slug in metricas_extra:
            metricas = metricas_extra[eq_slug]
        ret_exc = metricas.get("exc")
        puntos_detalle = {
            "ir": num(val("pts_ir")) or 0, "exc": num(val("pts_exc")) or 0,
            "sharpe": num(val("pts_sharpe")) or 0, "var95": num(val("pts_var95")) or 0,
            "mdd": num(val("pts_mdd")) or 0,
        }
        if not any(puntos_detalle.values()) and eq_slug in puntos_extra:
            puntos_detalle = puntos_extra[eq_slug]
        equipos.append({
            "id": eq_slug,
            "nombre": str(nombre).strip(),
            "posicion": int(num(val("posicion")) or 0),
            "puntos": round(num(val("puntos")) or 0.0, 2),
            "retRel": num(val("retRel")) if val("retRel") is not None else ret_exc,
            "delta": 0,  # se calcula después contra la semana anterior
            "metricas": metricas,
            "puntosDetalle": puntos_detalle,
            "miembros": [],
            "_ret": num(val("ret")),  # retorno acumulado del equipo (para historial)
        })
    return equipos


def _normalizar_linkedin(li):
    li = str(li or "").strip()
    if li and not li.startswith("http"):
        li = "https://" + li.lstrip("/")
    return li


def leer_inscripciones(ruta):
    """equipo (slug) -> [{nombre, linkedin}].

    Soporta dos formatos de Excel de inscripciones:
    - Largo: una fila por integrante (columnas equipo/nombre/linkedin).
    - Ancho (el real, "Copia de Inscripciones..."): una fila por equipo con
      columnas repetidas Líder/Int2/Int3 Nombre + LinkedIn (Carrera, Ingreso
      y Correo se leen pero NUNCA se guardan en el JSON — solo nombre y
      LinkedIn, que es lo único aprobado para publicar)."""
    filas = leer_hoja(ruta)
    encabezados = [normalizar(h) for h in filas[0]]

    # formato ancho: hay una columna "lider_nombre"
    if "lider_nombre" in encabezados:
        idx = {h: i for i, h in enumerate(encabezados)}
        equipo_i = idx.get("equipo")
        if equipo_i is None:
            print(f"  AVISO: inscripciones (formato ancho) sin columna 'equipo' "
                  f"({filas[0]}) — miembros quedarán vacíos.")
            return {}
        prefijos = ["lider", "int2", "int3", "int4", "int5"]
        miembros = {}
        for fila in filas[1:]:
            eq = fila[equipo_i] if equipo_i < len(fila) else None
            if not eq:
                continue
            integrantes = []
            for pre in prefijos:
                ni, li_i = idx.get(f"{pre}_nombre"), idx.get(f"{pre}_linkedin")
                if ni is None or ni >= len(fila):
                    continue
                nom = fila[ni]
                if not nom or not str(nom).strip():
                    continue
                li = _normalizar_linkedin(fila[li_i]) if li_i is not None and li_i < len(fila) else ""
                integrantes.append({"nombre": str(nom).strip(), "linkedin": li})
            if integrantes:
                miembros[slug(eq)] = integrantes
        return miembros

    # formato largo
    mapa = mapear_columnas(filas[0], ALIAS_INSC)
    if "equipo" not in mapa or "nombre" not in mapa:
        print(f"  AVISO: inscripciones sin columnas equipo/nombre reconocibles "
              f"({filas[0]}) — miembros quedarán vacíos.")
        return {}
    miembros = {}
    for fila in filas[1:]:
        def val(k):
            return fila[mapa[k]] if k in mapa and mapa[k] < len(fila) else None
        eq, nom = val("equipo"), val("nombre")
        if not eq or not nom:
            continue
        li = _normalizar_linkedin(val("linkedin"))
        miembros.setdefault(slug(eq), []).append({"nombre": str(nom).strip(), "linkedin": li})
    return miembros


def cargar_anterior():
    if SALIDA.exists():
        try:
            return json.loads(SALIDA.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError) as e:
            print(f"  AVISO: no pude leer el torneo.json anterior ({e}) — historial parte de cero.")
    return None


def integrar_historial(equipos, anterior, semana):
    """conserva el historial previo por id y agrega la semana actual; calcula delta."""
    previos = {}
    if anterior:
        for eq in anterior.get("equipos", []):
            previos[eq["id"]] = eq
    for eq in equipos:
        prev = previos.get(eq["id"])
        hist = list(prev.get("historial", [])) if prev else []
        hist = [h for h in hist if h.get("semana") != semana]  # re-ejecución de la misma semana
        hist.append({"semana": semana, "puntos": eq["puntos"], "posicion": eq["posicion"],
                     "ret": eq.pop("_ret", None)})
        hist.sort(key=lambda h: h["semana"])
        eq["historial"] = hist
        if prev and len(hist) >= 2:
            eq["delta"] = hist[-2]["posicion"] - eq["posicion"]  # + = sube
        if prev and not eq["miembros"]:
            eq["miembros"] = prev.get("miembros", [])


def demo():
    """genera un torneo.json de prueba (mismos 8 equipos del modo DEMO de la página)."""
    import random
    random.seed(2026)
    nombres = ["Compass Advisors", "Alpha Capital", "Bull & Bear", "FENIX Group",
               "Horizon Capital", "Delta Global", "Equinox Wealth", "Granite Partners"]
    equipos = []
    for i, n in enumerate(nombres):
        base = 100 - i * 12.5
        hist = [{"semana": s + 1,
                 "puntos": round(max(0, base - (7 - s) * 4 + random.uniform(-3, 3)), 2),
                 "posicion": i + 1,
                 "ret": round(.02 * (s + 1) * (1 - i * .06), 4)} for s in range(8)]
        equipos.append({
            "id": slug(n), "nombre": n, "posicion": i + 1,
            "puntos": hist[-1]["puntos"], "retRel": round(0.08 - i * 0.01, 4), "delta": 0,
            "metricas": {"ir": round(1 - i * .12, 2), "exc": round(.08 - i * .01, 4),
                         "sharpe": round(1.6 - i * .15, 2), "var95": round(-.01 - i * .002, 4),
                         "mdd": round(.05 + i * .017, 3)},
            "puntosDetalle": {"ir": round(30 - i * 3.7, 2), "exc": round(25 - i * 3.1, 2),
                              "sharpe": round(15 - i * 1.9, 2), "var95": round(15 - i * 1.9, 2),
                              "mdd": round(15 - i * 1.9, 2)},
            "miembros": [], "historial": hist,
        })
    acwi = [{"semana": s + 1, "ret": round(.01 * (s + 1) * .97, 4)} for s in range(8)]
    return {"semana": 8, "corte": "03 · JUL · 2026", "acwi": acwi, "equipos": equipos}


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    ap.add_argument("--excel", help="Excel oficial con hoja ranking_ordenado")
    ap.add_argument("--inscripciones", help="Excel de inscripciones (columna LinkedIn)")
    ap.add_argument("--semana", type=int, help="número de semana del corte (1-25)")
    ap.add_argument("--corte", help='fecha del corte, ej: "03 · JUL · 2026"')
    ap.add_argument("--acwi", type=float, help="retorno acumulado del ACWI esta semana, decimal (ej: 0.078)")
    ap.add_argument("--demo", action="store_true", help="escribe un torneo.json ficticio de prueba")
    ap.add_argument("--salida", default=str(SALIDA), help=f"ruta de salida (default {SALIDA})")
    args = ap.parse_args()

    salida = Path(args.salida)
    if args.demo:
        datos = demo()
    else:
        if not (args.excel and args.semana and args.corte):
            ap.error("--excel, --semana y --corte son obligatorios (o usa --demo)")
        equipos = leer_ranking(args.excel)
        if args.inscripciones:
            insc = leer_inscripciones(args.inscripciones)
            for eq in equipos:
                eq["miembros"] = insc.get(eq["id"], [])
        anterior = cargar_anterior()
        integrar_historial(equipos, anterior, args.semana)
        acwi = list((anterior or {}).get("acwi", []))
        acwi = [a for a in acwi if a.get("semana") != args.semana]
        if args.acwi is not None:
            acwi.append({"semana": args.semana, "ret": args.acwi})
        acwi.sort(key=lambda a: a["semana"])
        datos = {"semana": args.semana, "corte": args.corte, "acwi": acwi, "equipos": equipos}

    salida.parent.mkdir(parents=True, exist_ok=True)
    salida.write_text(json.dumps(datos, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"OK · {salida} · semana {datos['semana']} · {len(datos['equipos'])} equipos")
    print("La página torneo/index.html saldrá del modo DEMO automáticamente.")


if __name__ == "__main__":
    main()
